import datetime

from django.views.generic import View, TemplateView
from django.db import connection
from django.core.paginator import Paginator
from django.forms.models import model_to_dict
from django.http import JsonResponse

from django.contrib.auth.mixins import LoginRequiredMixin

from zipfile import BadZipfile

from .models import DecidirCard, DecidirPMC, GireModel, SIRModel, ConciliacionModel, OperationTypeModel, ResultModel
from .utils import Utils
from .filters import DecidirCardFilter, DecidirPMCFilter, GireFilter, SIRFilter, ConciliacionFilter, ResultFilter
import csv, codecs
from django.http import HttpResponse
import logging
import pandas as pd
import openpyxl
import xlwt
import django.utils.timezone as timezone

logger = logging.getLogger(__name__)

logger.setLevel(logging.WARNING)
handler = logging.FileHandler('log_file.log')
formatter = logging.Formatter('%(asctime)s : %(name)s  : %(funcName)s : %(levelname)s : %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)

class IndexView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/dashboard.html'

    def get(self, request, *args, **kwargs):
        decidir_count = DecidirCard.objects.all().count()
        gire_count = GireModel.objects.all().count()
        sir_count = SIRModel.objects.all().count()
        #decidir_visa = DecidirModel.objects.filter('tarjeta'=='visa')
        context = {'total_decidir':decidir_count,'total_gire':gire_count,'total_sir':sir_count, 'usuario': request.user}

        return self.render_to_response(context)


class DecidirCardDetailView(TemplateView):

    def get(self, request, *args, **kwargs):

        decidir_object = DecidirCard.objects.get(pk=kwargs['pk'])

        decidir_object_dict = model_to_dict(decidir_object)

        return JsonResponse(decidir_object_dict)

class DecidirPMCDetailView(TemplateView):

    def get(self, request, *args, **kwargs):

        decidir_object = DecidirPMC.objects.get(pk=kwargs['pk'])

        decidir_object_dict = model_to_dict(decidir_object)

        return JsonResponse(decidir_object_dict)

class DecidirCardView(LoginRequiredMixin, TemplateView):

    template_name = 'importaciones/decidir/decidir_card.html'
    
    def get(self, request, *args, **kwargs):

        try:
            responseExportar=''
            exportar = request.GET.get('exportar_hidden')
            queryset = DecidirCard.objects.all()
            filter = DecidirCardFilter(request.GET, queryset=queryset)

            paginator = Paginator(filter.qs, 10)

            page_number = int(request.GET.get('page') if request.GET.get('page') else 1)
            paginator = paginator.get_page(page_number)

            context = {
                'paginator': paginator,
                'page_number': page_number,
                'status_list': queryset.values('estado_cierre').distinct(),
                'card_list': queryset.values('tarjeta').distinct(),
                'query': Utils.get_query(request),
                'paginator_dict': Utils.get_paginator_dict(paginator, page_number),
            }


        except Exception as e:
            context = {
                'show_alert': True,
                'message': 'Ocurrió un error al obtener los resultados',
                'alert_type': 'danger'
            }

        if exportar == 'exportar':
            responseExportar=self.export_excel(request.GET, queryset=filter.qs)

            if responseExportar.status_code == 200:
                return responseExportar
            else:
                context = {
                    'paginator': paginator,
                    'page_number': page_number,
                    'status_list': queryset.values('estado_cierre').distinct(),
                    'card_list': queryset.values('tarjeta').distinct(),
                    'query': Utils.get_query(request),
                    'paginator_dict': Utils.get_paginator_dict(paginator, page_number),
                }

        return self.render_to_response(context)

    def export_excel(self, request, *args, **kwargs):

        try:

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="Decidir_file.xls"'
            queryset = DecidirCard.objects.all()
            filter = DecidirCardFilter(request, queryset=queryset)
            card = filter.qs.values_list('id_operacion', 'titular', 'fecha_original', 'estado_cierre', 'motivo', 'tarjeta', 'monto', 'lote','matched')

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('SIR')
            row_num = 0
            font_style = xlwt.XFStyle()
            font_style.font.bold=True

            columnas = ['id_operacion', 'titular', 'fecha_original', 'estado_cierre', 'motivo', 'tarjeta', 'monto', 'lote','matched']
            
            for col_num in range(len(columnas)):
                    ws.write(row_num, col_num, columnas[col_num], font_style)

            card = [[x.strftime("%d-%m-%Y") if isinstance(x, datetime.datetime) else x for x in row] for row in card ]
            
            for card_row in card:
                row_num += 1
                for col_num in range(len(card_row)):
                    ws.write(row_num, col_num, card_row[col_num], font_style)
#
                wb.save(response)
            return response
        except Exception as e:
            logger.exception(e)
            return HttpResponse('Error obteniendo el archivo excel', status=404)

class DecidirPMCView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/decidir/decidir_pmc.html'

    def get(self, request, *args, **kwargs):

        try:
            responseExportar=''
            exportar = request.GET.get('exportar_hidden')
            queryset = DecidirPMC.objects.all()
            filter = DecidirPMCFilter(request.GET, queryset=queryset)

            paginator = Paginator(filter.qs, 10)

            page_number = int(request.GET.get('page') if request.GET.get('page') else 1)
            paginator = paginator.get_page(page_number)

            context = {
                'paginator': paginator,
                'page_number': page_number,
                'status_list': queryset.values('estado').distinct(),
                'card_list': queryset.values('tarjeta').distinct(),
                'query': Utils.get_query(request),
                'paginator_dict': Utils.get_paginator_dict(paginator, page_number),
            }

            if exportar == 'exportar':
                responseExportar=self.export_excel(request.GET, queryset=filter.qs)

            if responseExportar.status_code == 200:
                return responseExportar
            else:
                context = {
                        'paginator': paginator,
                        'page_number': page_number,
                        'status_list': queryset.values('estado').distinct(),
                        'card_list': queryset.values('tarjeta').distinct(),
                        'query': Utils.get_query(request),
                        'paginator_dict': Utils.get_paginator_dict(paginator, page_number),
                }

        except Exception as e:
            context = {
                'show_alert': True,
                'message': 'Ocurrió un error al obtener los resultados',
                'alert_type': 'danger'
            }
            logger.exception(e)
        return self.render_to_response(context)

    def export_excel(self, request, *args, **kwargs):

            try:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename="PMC_file.xls"'
                queryset = DecidirPMC.objects.all()
                filter = DecidirPMCFilter(request, queryset=queryset)
                pmc = filter.qs.values_list('id_operacion', 'titular', 'fecha_original', 'estado_cierre', 'motivo', 'monto', 'lote','matched')

                wb = xlwt.Workbook(encoding='utf-8')
                ws = wb.add_sheet('PMC')
                row_num = 0
                font_style = xlwt.XFStyle()
                font_style.font.bold=True

                columnas = ['id_operacion', 'titular', 'fecha_original', 'estado_cierre', 'motivo', 'monto', 'lote','matched']
    
                for col_num in range(len(columnas)):
                        ws.write(row_num, col_num, columnas[col_num], font_style)

                pmc = [[x.strftime("%d-%m-%Y") if isinstance(x, datetime.datetime) else x for x in row] for row in card ]
                
                for pmc_row in card:
                    row_num += 1
                    for col_num in range(len(pmc_row)):
                        ws.write(row_num, col_num, pmc_row[col_num], font_style)
    #
                    wb.save(response)
                return response
            except Exception as e:
                logger.exception(e)
                return HttpResponse('Error obteniendo el archivo excel', status=404)

class DecidirCardUploadView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/decidir/decidir_upload_card.html'

    tarjetas = {
        'Visa': {
            'lote_required': 'true'
        },
        'Amex': {
            'lote_required': 'true'
        },
        'MasterCard': {
            'lote_required': 'false'
        }
    }

    def get_context_data(self):

        context = {
            'tarjetas': self.tarjetas
        }

        return context

    def post(self, request, *args, **kwargs):

        try:

            if 'subir' in request.POST:

                self.validate(request)

                Utils.validate_empty_file(request)

                results = DecidirCard.objects.insert_decidir_excel_card_values(request)

                if results['i'] > 0:
                    context = {'show_alert': True, 'alert_type': 'warning', 'message': (('Se van a importar %s registros. Registros duplicados: %s') % (str(results['i']), str(results['duplicate']))), 'timestamp': str(results['timestamp'])}
                else:
                    if results['duplicate'] > 0:
                        context = {'show_alert': True, 'alert_type': 'danger', 'message': 'Hay %s registros duplicados. No hay registros nuevos' % str(results['duplicate'])}
                    else:
                        context = {'show_alert': True, 'alert_type': 'danger', 'message': 'No hay registros para guardar'}

            elif 'confirmar' in request.POST:

                DecidirCard.objects.confirm_upload_card(request.POST['timestamp'])

                context = {'show_alert': True, 'alert_type': 'success', 'message': 'Archivo subido', 'lote': ''}
            elif 'cancelar' in request.POST:

                DecidirCard.objects.cancel_upload_card(request.POST['timestamp'])

                context = {'show_alert': True, 'alert_type': 'success',
                           'message': 'Operación cancelada'}
        except Exception as e:
            context = {'show_alert': True, 'alert_type': 'danger', 'message': str(e)}

        context['tarjetas'] = self.tarjetas

        return self.render_to_response(context)

    def validate(self, request):

        tarjetas = {key:value for (key,value) in self.tarjetas.items() if value['lote_required'] == 'true'}

        if request.POST['tarjeta_select'] in tarjetas:
            self.validate_empty_field(request.POST['lote'], 'lote')

    def validate_empty_field(self, val, name):
        if not val:
            raise Exception('Debe ingresar el campo ' + name)

class DecidirPMCUploadView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/decidir/decidir_upload_pmc.html'

    def post(self, request, *args, **kwargs):

        try:

            if 'subir' in request.POST:

                Utils.validate_empty_file(request)

                results = DecidirPMC.objects.insert_decidir_excel_pmc_values(request)

                if results['i'] > 0:
                    context = {'show_alert': True, 'alert_type': 'warning', 'message': (('Se van a importar %s registros. Registros duplicados: %s') % (str(results['i']), str(results['duplicate']))), 'timestamp': str(results['timestamp'])}
                else:
                    if results['duplicate'] > 0:
                        context = {'show_alert': True, 'alert_type': 'danger', 'message': 'Hay %s registros duplicados. No hay registros nuevos' % str(results['duplicate'])}
                    else:
                        context = {'show_alert': True, 'alert_type': 'danger', 'message': 'No hay registros para guardar'}

            elif 'confirmar' in request.POST:

                DecidirPMC.objects.confirm_upload_pmc(request.POST['timestamp'])

                context = {'show_alert': True, 'alert_type': 'success', 'message': 'Archivo subido', 'lote': ''}
            elif 'cancelar' in request.POST:

                DecidirPMC.objects.cancel_upload_pmc(request.POST['timestamp'])

                context = {'show_alert': True, 'alert_type': 'success',
                           'message': 'Operación cancelada'}
        except Exception as e:
            context = {'show_alert': True, 'alert_type': 'danger', 'message': str(e)}

        return self.render_to_response(context)

class GireView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/gire/gire.html'

    def get(self, request, *args, **kwargs):
        try:
            responseExportar=''
            exportar = request.GET.get('exportar_hidden')
            queryset = GireModel.objects.all()
            filter = GireFilter(request.GET, queryset=queryset)

            paginator = Paginator(filter.qs, 10)

            page_number = int(request.GET.get('page') if request.GET.get('page') else 1)
            paginator = paginator.get_page(page_number)

            context = {
                'paginator': paginator,
                'page_number': page_number,
                'query': Utils.get_query(request),
                'paginator_dict': Utils.get_paginator_dict(paginator, page_number)
            }
        except Exception as e:
            context = {
                'show_alert': True,
                'message': 'Ocurrió un error al obtener los resultados',
                'alert_type': 'danger'
            }
            logger.exception(e)

        if exportar == 'exportar':
            responseExportar=self.export_excel(request.GET, queryset=filter.qs)

            if responseExportar.status_code == 200:
                return responseExportar
            else:
                context = {
                    'paginator': paginator,
                    'page_number': page_number,
                    'query': Utils.get_query(request),
                    'paginator_dict': Utils.get_paginator_dict(paginator, page_number)
                }

        
        return self.render_to_response(context)

    def export_excel(self, request, *args, **kwargs):

        try:

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="GIRE_file.xls"'
            queryset = GireModel.objects.all()
            filter = GireFilter(request, queryset=queryset)
            gire = filter.qs.values_list('barcode', 'fecha', 'importe', 'archivo')

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('GIRE')
            row_num = 0
            font_style = xlwt.XFStyle()
            font_style.font.bold=True

            columnas = ['BARCODE','FECHA', 'IMPORTE','ARCHIVO']
            
            for col_num in range(len(columnas)):
                    ws.write(row_num, col_num, columnas[col_num], font_style)

            gire = [[x.strftime("%d-%m-%Y") if isinstance(x, datetime.datetime) else x for x in row] for row in gire ]
            
            for gire_row in gire:
                row_num += 1
                for col_num in range(len(gire_row)):
                    ws.write(row_num, col_num, gire_row[col_num], font_style)

            wb.save(response)
            return response
        except Exception as e:
            logger.exception(e)
            return HttpResponse('Error obteniendo el archivo excel', status=404)

class GireUploadView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/gire/gire_upload.html'

    def post(self, request, *args, **kwargs):

        try:

            if 'subir' in request.POST:

                Utils.validate_empty_file(request)

                results = GireModel.objects.insert_values(request)

                if results['i'] > 0:
                    context = {'show_alert': True, 'alert_type': 'warning', 'message': 'Se van a importar %s registros. Registros duplicados: %s' % (str(results['i']), str(results['duplicate'])), 'timestamp': results['timestamp']}
                else:
                    if results['duplicate'] > 0:
                        context = {'show_alert': True, 'alert_type': 'danger', 'message': 'Hay %s registros duplicados. No hay registros nuevos' % str(results['duplicate'])}
                    else:
                        context = {'show_alert': True, 'alert_type': 'danger', 'message': 'No hay registros para guardar'}

            elif 'confirmar' in request.POST:

                GireModel.objects.confirm_upload(request.POST['timestamp'])

                context = {'show_alert': True, 'alert_type': 'success',
                           'message': 'Archivo subido'}
            elif 'cancelar' in request.POST:

                GireModel.objects.cancel_upload(request.POST['timestamp'])

                context = {'show_alert': True, 'alert_type': 'success',
                           'message': 'Operación cancelada'}

        except Exception as e:
            context = {'show_alert': True, 'alert_type': 'danger', 'message': str(e)}

        return self.render_to_response(context)

class SirDetailView(TemplateView):

    def get(self, request, *args, **kwargs):
        sir_object = SIRModel.objects.get(pk=kwargs['pk'])

        sir_dict = model_to_dict(sir_object)


        return JsonResponse(sir_dict)

class SIRView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/sir/sir.html'
    def get_context_data(TemplateView, **kwargs):
        #get data from default method
        context = super().get_context_data(**kwargs)

    def get(self, request, *args, **kwargs):
        try:
            responseExportar=''
            exportar = request.GET.get('exportar_hidden')

            queryset = SIRModel.objects.all().order_by('id')
            filter = SIRFilter(request.GET, queryset=queryset)

            paginator = Paginator(filter.qs, 10)

            page_number = int(request.GET.get('page') if request.GET.get('page') else 1)
            paginator = paginator.get_page(page_number)

            context = {
                'paginator': paginator,
                'page_number': page_number,
                'query': Utils.get_query(request),
                'paginator_dict': Utils.get_paginator_dict(paginator, page_number),
                'medios_de_pago': queryset.values('medio_de_pago').distinct().order_by('medio_de_pago'),
                'tipos_de_tramite': OperationTypeModel.objects.filter(id__in=queryset.values('codigo_de_tramite_id').distinct())
            }
        except Exception as e:
            context = {
                'show_alert': True,
                'message': 'Ocurrió un error al obtener los resultados',
                'alert_type': 'danger'
            }
            logger.exception(e)

        if exportar == 'exportar':
            responseExportar=self.export_sir_excel(request.GET, queryset=filter.qs)

            if responseExportar.status_code == 200:
                return responseExportar
            else:
                context = {
                    'paginator': paginator,
                    'page_number': page_number,
                    'query': Utils.get_query(request),
                    'paginator_dict': Utils.get_paginator_dict(paginator, page_number),
                    'medios_de_pago': queryset.values('medio_de_pago').distinct().order_by('medio_de_pago'),
                    'tipos_de_tramite': OperationTypeModel.objects.filter(id__in=queryset.values('codigo_de_tramite_id').distinct())
                }

        return self.render_to_response(context)

    def export_sir_excel(self, request, *args, **kwargs):

        try:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="SIR_file.xls"'
            queryset = SIRModel.objects.all()
            filter = SIRFilter(request, queryset=queryset)
            sir = filter.qs.values_list('transaccion','medio_de_pago', 'monto', 'fecha_de_pago', 'id_tad','dominio','cantidad_de_dominios','numero_de_factura','estado_en_sir','codigo_de_tramite__name')

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('SIR')
            row_num = 0
            font_style = xlwt.XFStyle()
            font_style.font.bold=True

            columnas = ['TRANSACCION','MEDIO DE PAGO', 'MONTO','FECHA DE PAGO','ID TAD','DOMINIO','CANTIDAD DE DOMINIOS','FACTURA','ESTADO EN SIR','TRAMITE']
            
            for col_num in range(len(columnas)):
                    ws.write(row_num, col_num, columnas[col_num], font_style)

            sir = [[x.strftime("%d-%m-%Y") if isinstance(x, datetime.datetime) else x for x in row] for row in sir ]
            
            for sir_row in sir:
                row_num += 1
                for col_num in range(len(sir_row)):
                    ws.write(row_num, col_num, sir_row[col_num], font_style)

                wb.save(response)
            return response
        except Exception as e:
            logger.exception(e)
            return HttpResponse('Error obteniendo el archivo excel', status=404)

class SIRUploadView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/sir/sir_upload.html'

    def post(self, request, *args, **kwargs):

        try:

            if 'subir' in request.POST:
                Utils.validate_empty_file(request)
                
                results = SIRModel.objects.insert_values(request)

                if results['i'] > 0:
                    context = {'show_alert': True, 'alert_type': 'warning', 'message': 'Se van a importar %s registros. Registros duplicados: %s' % (str(results['i']), str(results['duplicate'])), 'timestamp': results['timestamp']}
                else:
                    if results['duplicate'] > 0:
                        context = {'show_alert': True, 'alert_type': 'info', 'message': '%s registros guardados. No hay registros nuevos' % str(results['duplicate'])}
                    else:
                        context = {'show_alert': True, 'alert_type': 'danger', 'message': 'No hay registros para guardar'}


            elif 'confirmar' in request.POST:

                SIRModel.objects.confirm_upload(request.POST['timestamp'])

                context = {'show_alert': True, 'alert_type': 'success', 'message': 'Archivo subido'}
            elif 'cancelar' in request.POST:

                SIRModel.objects.cancel_upload(request.POST['timestamp'])

                context = {'show_alert': True, 'alert_type': 'success', 'message': 'Operación cancelada'}

        except BadZipfile:
            context = {'show_alert': True, 'alert_type': 'danger', 'message': 'Formato de archivo no válido'}
        except Exception as e:
            context = {'show_alert': True, 'alert_type': 'danger', 'message': str(e)}

        return self.render_to_response(context)

class ConciliacionesDetailView(View):

    def get(self, request, *args, **kwargs):

        object = ConciliacionModel.objects.get(pk=kwargs['pk'])
        object_sir = SIRModel.objects.get(transaccion=object.transaccion)
        
        oo = model_to_dict(object)
        oo['tramite'] = object.codigo_de_tramite.name
        oo['numero_de_factura'] = object_sir.numero_de_factura
        
        if object_sir.estado_en_sir:
            estado_en_sir = 'PAGO'
        else:
            estado_en_sir = 'SIN PAGO'
        oo['estado_en_sir'] = estado_en_sir

        return JsonResponse(oo)

class ConciliacionesView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/conciliaciones/conciliaciones.html'
    
    def get_context_data(TemplateView, **kwargs):
        #get data from default method
        context = super().get_context_data(**kwargs)
    
    def get(self, request, *args, **kwargs):
        try:
            responseExportar=''
            exportar = request.GET.get('exportar_hidden')
            
            queryset = ConciliacionModel.objects.all()
            filter = ConciliacionFilter(request.GET, queryset=queryset)

            paginator = Paginator(filter.qs, 10)

            page_number = int(request.GET.get('page') if request.GET.get('page') else 1)
            paginator = paginator.get_page(page_number)

            context = {
                'paginator': paginator,
                'page_number': page_number,
                'query': Utils.get_query(request),
                'paginator_dict': Utils.get_paginator_dict(paginator, page_number),
                'medios_de_pago': queryset.values('medio_de_pago').distinct().order_by('medio_de_pago'),
                'codigos_de_tramite_enum': OperationTypeModel.objects.filter(id__in=queryset.values('codigo_de_tramite')),
            }
        except Exception as e:
            context = {
                'show_alert': True,
                'message': 'Ocurrió un error al obtener los resultados',
                'alert_type': 'danger'
            }
            logger.exception(e)

        if exportar == 'exportar':
            #responseExportar=self.export_conciliaciones_csv(request.GET, queryset=filter.qs)
            responseExportar=self.export_conciliaciones_excel(request.GET, queryset=filter.qs)
            if responseExportar.status_code == 200:
                return responseExportar
            else:
                context = {
                'show_alert': True,
                'message': 'No se pudo crear el archivo de exportación',
                'alert_type': 'danger',
                'paginator': paginator,
                'page_number': page_number,
                'query': Utils.get_query(request),
                'paginator_dict': Utils.get_paginator_dict(paginator, page_number),
                'medios_de_pago': queryset.values('medio_de_pago').distinct().order_by('medio_de_pago'),
                'codigos_de_tramite_enum': OperationTypeModel.objects.all(),
            }

        return self.render_to_response(context)

    def export_conciliaciones_excel(self, request, *args, **kwargs):

        try:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="concilicaciones_file.xls"'
            queryset =  ConciliacionModel.objects.all()
            filter = ConciliacionFilter(request, queryset=queryset)
            conciliaciones = filter.qs.values_list('transaccion','medio_de_pago', 'monto', 'fecha_de_pago', 'id_tad','dominios','numero_de_factura','estado_en_sir','codigo_de_tramite__name')

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('CONCILIACIONES')
            row_num = 0
            font_style = xlwt.XFStyle()
            font_style.font.bold=True

            columnas = ['TRANSACCION','MEDIO DE PAGO', 'MONTO','FECHA DE PAGO','ID TAD','DOMINIOS','FACTURA','ESTADO EN SIR','TRAMITE']
            
            for col_num in range(len(columnas)):
                    ws.write(row_num, col_num, columnas[col_num], font_style)

            conciliaciones = [[x.strftime("%d-%m-%Y") if isinstance(x, datetime.datetime) else x for x in row] for row in conciliaciones ]
            
            for conciliacion_row in conciliaciones:
                row_num += 1
                for col_num in range(len(conciliacion_row)):
                    ws.write(row_num, col_num, conciliacion_row[col_num], font_style)

                wb.save(response)
            return response
        except Exception as e:
            logger.exception(e)
            return HttpResponse('Error obteniendo el archivo excel', status=404)

    def export_conciliaciones_csv(self, request, *args, **kwargs):

        try:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="conciliaciones_file.csv"'
            response.write(codecs.BOM_UTF8)
            queryset = ConciliacionModel.objects.all()
            #total = queryset.count()
            filter = ConciliacionFilter(request, queryset=queryset)

            nombre = 'conciliaciones_file.csv'           
           
            conciliaciones = filter.qs.values_list('transaccion','medio_de_pago', 'monto','id_tad','fecha_de_pago','dominios','numero_de_factura','estado_en_sir','codigo_de_tramite__name')
            
            writer = csv.writer(response)
            
            for conciliacion in conciliaciones:
                writer.writerow(conciliacion)

            return response
        except Exception as e:
            logger.exception(e)
            return HttpResponse('Error obteniendo el archivo csv', status=404)

class Dashboard(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/dashboard.html'

    def get(self, request, *args, **kwargs):
       registrosSIRImpagos = None
       decidir_count_card = DecidirCard.objects.all().count()
       decidir_count_card_visa = DecidirCard.objects.filter(tarjeta='Visa').count()
       decidir_count_card_amex = DecidirCard.objects.filter(tarjeta='Amex').count()
       decidir_count_card_master = DecidirCard.objects.filter(tarjeta='Mastercard').count()
       decidir_count_card = DecidirCard.objects.all().count()
       decidir_count_pmc = DecidirPMC.objects.all().count()
       decidir_count = decidir_count_card + decidir_count_pmc
       gire_count = GireModel.objects.all().count()
       sir_count = SIRModel.objects.all().count()
       
       countMatchedSIR = SIRModel.objects.filter(matched='1').count()
       countUNMatchedSIR = SIRModel.objects.filter(matched='0').count()
       countMatchedSIRSinFactura = SIRModel.objects.filter(matched='1').filter(numero_de_factura=None).count()

       countUnmatchedDecidirCard = DecidirCard.objects.filter(matched='0').count()
       countUnmatchedDecidirPMC = DecidirPMC.objects.filter(matched='0').count()
       countUnmatchedGIRE = GireModel.objects.filter(matched='0').count()

       try:
            with connection.cursor() as cursor:
                registrosSIRImpagos = cursor.execute("select * from raw_data_sir where estado_en_sir=false")
       except Exception as e:
           registrosSIRImpagos = 'No se pudo obtener la cantidad de registros impagos en SIR.'

       context = {'total_decidir': decidir_count_card,
                  'total_decidir_card_visa': decidir_count_card_visa,
                  'total_decidir_card_amex': decidir_count_card_amex,
                  'total_decidir_card_master': decidir_count_card_master,
                  'total_decidir_pmc': decidir_count_pmc,
                  'total_gire': gire_count,
                  'total_sir': sir_count,
                  'total_matched_sir': countMatchedSIR,
                  'total_unmatched_sir': countUNMatchedSIR,
                  'usuario': request.user,
                  'registrosSIRImpagos': registrosSIRImpagos,
                  'total_unmatched_decidir_card': countUnmatchedDecidirCard,
                  'total_unmatched_decidir_pmc': countUnmatchedDecidirPMC,
                  'total_unmatched_gire': countUnmatchedGIRE,
                  'total_matched_sin_factura':countMatchedSIRSinFactura,
                  }

       return self.render_to_response(context)

class ResultsView(LoginRequiredMixin, TemplateView):
    template_name = 'importaciones/conciliaciones/results.html'

    def get(self, request, *args, **kwargs):

        if request.GET:
            queryset = ResultModel.objects.all()
            filter = ResultFilter(request.GET, queryset=queryset)
            queryset = filter.qs
        else:
            queryset = ResultModel.objects.filter(year=datetime.date.today().year, month=datetime.date.today().month)


        context = {
            'results': queryset,
            'months': self.get_months(),
            'years': self.get_years()
        }

        return self.render_to_response(context)

    def get_months(self):


        months_dict = {
            '1': 'Enero',
            '2': 'Febrero',
            '3': 'Marzo',
            '4': 'Abril',
            '5': 'Mayo',
            '6': 'Junio',
            '7': 'Julio',
            '8': 'Agosto',
            '9': 'Septiembre',
            '10': 'Octubre',
            '11': 'Noviembre',
            '12': 'Diciembre',
        }

        return months_dict

    def get_years(self):


        years = []
        current_year = datetime.date.today().year
        years.append(current_year)

        for i in range(1,10):
            years.append(current_year - i)

        return years